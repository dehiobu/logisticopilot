import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from datetime import datetime

def export_manifest_to_excel(df, summary_text, compliance_notes=None):
    """
    Exports a logistics manifest DataFrame, AI-generated summary, and optional compliance notes
    into a multi-sheet Excel workbook. It also includes example Copilot prompts.

    Args:
        df (pd.DataFrame): The pandas DataFrame containing the logistics manifest data.
        summary_text (str): A string containing the AI-generated summary of the manifest.
        compliance_notes (list, optional): A list of strings, each representing a compliance issue.
                                           If None or empty, the "Compliance" sheet is not created.

    Returns:
        str: The full path to the generated Excel file.
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Manifest"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)

    ws1["F1"] = "Delay Risk"
    ws1["F2"] = "=IF(TODAY()-D2>5, \"⚠️ Delay Risk\", \"\")"
    ws1["G1"] = "Delivered?"
    ws1["G2"] = "=IF(C2=\"Delivered\", \"✅\", \"\")"

    ws2 = wb.create_sheet("AI Summary")
    ws2.append(["Generated Summary"])
    for line in summary_text.split("\n"):
        ws2.append([line])

    if compliance_notes:
        ws3 = wb.create_sheet("Compliance")
        ws3.append(["Compliance Issues"])
        for note in compliance_notes:
            ws3.append([note])

    prompts = [
        ["Prompt Example", "Description"],
        ["Highlight all delayed shipments", "Uses conditional formatting to tag 'Delayed' rows"],
        ["Summarise carrier performance", "Group by Carrier and count statuses"],
        ["Group by destination and count delays", "Build pivot table using Status column"],
        ["Add delay risk indicator", "Use formula: =IF(TODAY()-[@ShipDate]>5, '⚠️ Delay Risk', '')"],
        ["Mark delivered shipments", "Use formula: =IF([@Status]='Delivered', '✅', '')"]
    ]
    ws4 = wb.create_sheet("Copilot Prompts")
    for row in prompts:
        ws4.append(row)
    ws4["A1"].font = ws4["B1"].font = Font(bold=True)

    filename = f"/tmp/export_logibot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename
